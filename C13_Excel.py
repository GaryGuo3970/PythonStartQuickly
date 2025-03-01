import logging
import openpyxl

logging.basicConfig(level=logging.DEBUG,format= '%(asctime)s %(levelname)s %(message)s')

#加载excel
wb = openpyxl.load_workbook("excel/Demo.xlsx")
logging.debug(type(wb))

#读取sheet页
wbsheets = wb.sheetnames
for n in wbsheets:
    logging.debug(f'正在读取sheet {n}')
    currentSheet=wb[n]
    logging.debug(currentSheet.title)
    logging.debug(wb.active)  #当前活动的sheet

    #读取sheet内容
    logging.debug(f'读取sheet {currentSheet.title} 单元格的值：{currentSheet['A1'].value}')

    #读取
    for i in range(1,10):
        for j in range(1,3):
            logging.debug(f'行{i} 列{j} {currentSheet.cell(row=i,column=j).value}')


