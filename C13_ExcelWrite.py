import openpyxl
import openpyxl.workbook
from openpyxl.styles import Font

#新建一个excel对象
wb = openpyxl.Workbook()
#开始一个sheet
wb.sheetnames
sheet=wb.active
sheet.title ='FirstSheet'

#新建第二个sheet
wb.create_sheet(index=1,title='第二个sheet')
#新建第三个sheet
wb.create_sheet(index=2,title='第三个sheet')
#第二个第三之间插入一个sheet
wb.create_sheet(index=2,title='插入sheet')
print(wb.sheetnames)

#更新sheet
currentSheet=wb['第二个sheet']
currentSheet['A1'] = '行1列1'
currentSheet['A2'] = '行2列1'
currentSheet['B2'] = '行2列2'

#设置单元格格式
italic24Font = Font(size=24,italic=True,bold=True,color='FF0000')
currentSheet['A2'].font = italic24Font
#设置列宽
currentSheet.column_dimensions['A'].width=20

wb.save('excel/创建的excel.xlsx')


